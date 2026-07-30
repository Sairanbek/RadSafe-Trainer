import re
import json
from pathlib import Path
from openpyxl import load_workbook

from database.database import init_db, get_connection

QUESTIONS_FILE = Path(__file__).resolve().parent.parent / "questions" / "radiation_safety" / "RST_Вопросы_варианты_разделы.xlsx"
DATA_DIR = Path(__file__).resolve().parent / "data"

SEC_RE = re.compile(r'^РАЗДЕЛ\s+\d+\.\s*(.*?)\s*\(\d+\s*вопрос')
Q_RE = re.compile(r'^Вопрос\s+\d+\.\s*(.*)')
OPT_RE = re.compile(r'^([A-E])\)\s*(.*)')


def migrate_questions():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) as c FROM questions")
    if cur.fetchone()["c"] > 0:
        print("Вопросы уже есть в базе, пропускаю.")
        conn.close()
        return

    wb = load_workbook(QUESTIONS_FILE, data_only=True)
    ws = wb["Вопросы с вариантами"]

    current_section = "Без раздела"
    current_question_text = None
    current_options = []
    qid = 1
    count = 0

    def flush():
        nonlocal qid, current_question_text, current_options, count
        if current_question_text and len(current_options) == 5:
            correct_idx = next((i for i, o in enumerate(current_options) if "✓" in o), None)
            if correct_idx is not None:
                clean_opts = [o.replace("✓", "").strip() for o in current_options]
                correct = clean_opts.pop(correct_idx)
                cur.execute(
                    "INSERT INTO questions (id, section, question, answer, wrong1, wrong2, wrong3, wrong4) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (qid, current_section, current_question_text.strip(), correct,
                     clean_opts[0], clean_opts[1], clean_opts[2], clean_opts[3])
                )
                qid += 1
                count += 1
        current_question_text = None
        current_options = []

    for row in ws.iter_rows(values_only=True):
        cell = row[0]
        if not cell or not str(cell).strip():
            continue
        line = str(cell).strip()

        m = SEC_RE.match(line)
        if m:
            current_section = m.group(1).strip()
            continue

        m = Q_RE.match(line)
        if m:
            flush()
            current_question_text = m.group(1)
            continue

        m = OPT_RE.match(line)
        if m:
            current_options.append(m.group(2))
            continue

    flush()
    conn.commit()
    conn.close()
    print(f"Перенесено вопросов: {count}")


def migrate_users():
    path = DATA_DIR / "users.json"
    if not path.exists():
        print("users.json не найден, пропускаю.")
        return
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    conn = get_connection()
    cur = conn.cursor()
    for user_id, info in data.items():
        cur.execute(
            "INSERT OR REPLACE INTO users (user_id, first_name, username, first_seen, last_seen, visits) VALUES (?, ?, ?, ?, ?, ?)",
            (int(user_id), info.get("first_name", ""), info.get("username", ""),
             info.get("first_seen"), info.get("last_seen"), info.get("visits", 0))
        )
    conn.commit()
    conn.close()
    print(f"Перенесено пользователей: {len(data)}")


def migrate_stats():
    path = DATA_DIR / "stats.json"
    if not path.exists():
        print("stats.json не найден, пропускаю.")
        return
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    conn = get_connection()
    cur = conn.cursor()
    count = 0
    for user_id, sections in data.items():
        for section, s in sections.items():
            cur.execute(
                "INSERT OR REPLACE INTO stats (user_id, section, asked, correct) VALUES (?, ?, ?, ?)",
                (int(user_id), section, s.get("asked", 0), s.get("correct", 0))
            )
            count += 1
    conn.commit()
    conn.close()
    print(f"Перенесено строк статистики: {count}")


def migrate_mistakes():
    path = DATA_DIR / "mistakes.json"
    if not path.exists():
        print("mistakes.json не найден, пропускаю.")
        return
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    conn = get_connection()
    cur = conn.cursor()
    count = 0
    for user_id, ids in data.items():
        for qid in ids:
            cur.execute(
                "INSERT OR IGNORE INTO mistakes (user_id, question_id) VALUES (?, ?)",
                (int(user_id), int(qid))
            )
            count += 1
    conn.commit()
    conn.close()
    print(f"Перенесено ошибок: {count}")


def migrate_history():
    path = DATA_DIR / "history.json"
    if not path.exists():
        print("history.json не найден, пропускаю.")
        return
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    conn = get_connection()
    cur = conn.cursor()
    count = 0
    for user_id, entries in data.items():
        for e in entries:
            cur.execute(
                "INSERT INTO history (user_id, date, mode, section, total, correct, wrong, percent) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (int(user_id), e["date"], e["mode"], e["section"], e["total"], e["correct"], e["wrong"], e["percent"])
            )
            count += 1
    conn.commit()
    conn.close()
    print(f"Перенесено записей истории: {count}")


if __name__ == "__main__":
    init_db()
    migrate_questions()
    migrate_users()
    migrate_stats()
    migrate_mistakes()
    migrate_history()
    print("Миграция завершена!")