import re
from pathlib import Path
from openpyxl import load_workbook

from database.models import Question

SEC_RE = re.compile(r'^РАЗДЕЛ\s+\d+\.\s*(.*?)\s*\(\d+\s*вопрос')
Q_RE = re.compile(r'^Вопрос\s+\d+\.\s*(.*)')
OPT_RE = re.compile(r'^([A-E])\)\s*(.*)')


def load_questions(file_path=None):
    if file_path is None:
        project_root = Path(__file__).resolve().parents[2]
        file_path = (
            project_root / "questions" / "radiation_safety"
            / "RST_Вопросы_варианты_разделы.xlsx"
        )

    wb = load_workbook(file_path, data_only=True)
    ws = wb["Вопросы с вариантами"]

    questions = []
    current_section = "Без раздела"
    current_question_text = None
    current_options = []
    qid = 1

    def flush():
        nonlocal qid, current_question_text, current_options
        if current_question_text and len(current_options) == 5:
            correct_idx = next((i for i, o in enumerate(current_options) if "✓" in o), None)
            if correct_idx is not None:
                clean_opts = [o.replace("✓", "").strip() for o in current_options]
                correct = clean_opts.pop(correct_idx)
                questions.append(
                    Question(
                        id=qid,
                        section=current_section,
                        question=current_question_text.strip(),
                        answer=correct,
                        wrong_answers=clean_opts
                    )
                )
                qid += 1
        current_question_text = None
        current_options = []

    for row in ws.iter_rows(values_only=True):
        cell = row[0]
        if not cell or not str(cell).strip():
            continue
        line = str(cell).strip()

        m = SEC_RE.match(line)
        if m:
            current_section = m.group(1).strip()
            continue

        m = Q_RE.match(line)
        if m:
            flush()
            current_question_text = m.group(1)
            continue

        m = OPT_RE.match(line)
        if m:
            current_options.append(m.group(2))
            continue

    flush()
    print(f"Загружено вопросов: {len(questions)}")
    return questions


def get_sections(questions):
    seen = []
    for q in questions:
        if q.section not in seen:
            seen.append(q.section)
    return seen